from openpyxl import *
from openpyxl.styles import PatternFill
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule

wb = load_workbook("reviews-sample.xlsx")
sheet = wb["amazon_reviews_us_Watches_v1_00"]

red_background = PatternFill(fgColor="00FF0000")
diff_style = DifferentialStyle(fill=red_background)
rule = Rule(type="expression", dxf=diff_style)
rule.formula = ["$H1<3"]
sheet.conditional_formatting.add("A1:O100", rule)

wb.save("reviews-sample.xlsx")