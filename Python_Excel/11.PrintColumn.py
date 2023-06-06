# In this program we are going to print only column values.

from openpyxl import *

# Load workbook, select worksheet, count max number of rows in worksheet.
wb = load_workbook("sample.xlsx")
sheet = wb["Student"]
rowCount = sheet.max_row

# Loop through all the rows
# Note: Change column number in line number 15 to see the changes in output.

for i in range(1, rowCount + 1):
    # Iterating over all the rows of first column
    rowValue = sheet.cell(row=i, column=1)
    print(rowValue.value)