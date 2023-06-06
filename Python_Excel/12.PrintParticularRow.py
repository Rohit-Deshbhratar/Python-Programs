# In this program we are going to print particular row from data.

from openpyxl import *

# Load workbook, select worksheet, count max number of columns in worksheet.
wb = load_workbook("sample.xlsx")
sheet = wb["Student"]
colCount = sheet.max_column

# Loop through all the rows till find matching row
# Note: Change row number in line number 15 to see the changes in output.

for i in range(1, colCount + 1):
    #Row number remain same, but column number incremented by 1.
    rowValue = sheet.cell(row=4, column=i)
    print(rowValue.value, end = " ")