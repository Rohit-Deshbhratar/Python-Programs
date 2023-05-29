# In this program we are reading excel sheet names using ".sheetnames" and "workbook.active"

from openpyxl import load_workbook

workbook = load_workbook(filename="reviews-sample.xlsx")
print(workbook.sheetnames)

sheet = workbook.active
print(sheet)