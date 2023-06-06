# Delete existing worksheet from workbook
# In this program we are going to create a worksheet and delete that worksheet

from openpyxl import *

# Load workbook and print sheet names from workbooks
wb = load_workbook("sample.xlsx")
print(f"Exixting sheets in curent workbook:\n{wb.sheetnames}")

# Create new worksheet in current workbook
wb.create_sheet(index=4, title="To Be Deleted")
wb.save("sample.xlsx")
print(f"\nAfter adding new sheet in curent workbook:\n{wb.sheetnames}")

# Deleting a sheet
sheet = wb["To Be Deleted"]
wb.remove(sheet)
print(f"\nAfter deleting sheet in curent workbook:\n{wb.sheetnames}")
wb.save("sample.xlsx")