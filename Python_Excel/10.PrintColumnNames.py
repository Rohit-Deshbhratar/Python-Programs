# In this program we are going to print only column names.

import openpyxl

# Load workbook, select worksheet, get max column count
wb = openpyxl.load_workbook("sample.xlsx")
sheet = wb["Student"]
column = sheet.max_column

# Loop through column
# Row number will remain same i.e. 1, but column values are incremented by 1.
for i in range(1, column + 1):
    # Iterating over first row and printing column values.
    colValue = sheet.cell(row=1, column=i)
    print(colValue.value)