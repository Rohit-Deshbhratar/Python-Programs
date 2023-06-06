# In this program we are going to read all the data from particular excel sheet.

from openpyxl import *

# Load workbook and select worksheet
wb = load_workbook("sample.xlsx")
sheet = wb["Student"]

# Count max rows and columns in excel worksheet
rowCount = sheet.max_row
colCount = sheet.max_column

# Loop through excel worksheet  and print data
# ".value" is used to print values from particular cell 
for i in range(1, rowCount + 1):
    for j in range(1, colCount + 1):
        data = sheet.cell(row=i, column=j).value
        print(data, end="           ")
    print("\n")