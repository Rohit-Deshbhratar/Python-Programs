# In this program we are writing data to excel file.
# Before running this program make sure that the information 
# entered into excel file should be deleted 
# i.e. delete only row 6 from both "Employee" and "DEPT" sheet
# OR enter new data

from openpyxl import load_workbook
# Loading workbook
wb = load_workbook("sample.xlsx")
# Selecting sheet in which we want to enter data
sheet = wb["Employee"]

# First way to insert data into excel file
sheet['A6'] = 105
sheet['B6'] = "Viraj"
sheet['C6'] = "Wankar"
sheet['D6'] = "PR"
sheet['E6'] = 90000

# Second way to insert data into file
sheet = wb["DEPT"]
sheet.cell(row=6, column=1).value = 50
sheet.cell(row=6, column=2).value = "PR"

# Saving workbook
wb.save("sample.xlsx")