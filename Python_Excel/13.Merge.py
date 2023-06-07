# In this program we are going to merge and umnerge cells in excel by using a program.
# There are two ways to merge and unmerge cells using "openpyxl".
# Try to uncomment code at line number 15 and 21 to see unmerge effects.

from openpyxl import * 

# Load workbook, select worksheet
wb = load_workbook("sample.xlsx")
sheet = wb["Merge"]

# First way to merge cells
sheet.merge_cells('A3:E3')

# First way to unmerge cells
#sheet.unmerge_cells('A3:E3')

# Second way to merge cells
sheet.merge_cells(start_row=9,start_column=6,end_row=15,end_column=10)

# Second way to unmerge cells
#sheet.unmerge_cells(start_row=9,start_column=6,end_row=15,end_column=10)
wb.save('sample.xlsx')  

